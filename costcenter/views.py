from datetime import datetime, timedelta
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.contrib import messages
from django.urls import reverse

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_JUSTIFY

from costcenter.models import Consumo, Transaction, Cards, Distribution, Acreditaciones, UserProfile
from costcenter.forms import TransactionForm, MesesForm, ConsumoForm

import json

def create_transaction(request):
    if request.method == 'POST':
        # Captura los datos del formulario
        from_account_number = request.POST.get('from_account')
        to_account_number = request.POST.get('to_account')
        amount = float(request.POST.get('amount', 0))
        
        # Validación de que la cuenta de origen esté seleccionada
        if not from_account_number:
            messages.error(request, 'Por favor, seleccione una cuenta de origen.')
            return render(request, 'costcenter/RealizarTransferencias.html')  # Renderiza la vista de formulario de nuevo
        
        try:
            # Obtén las tarjetas basadas en los números de cuenta
            from_card = Cards.objects.get(card_number=from_account_number, user=request.user)
            to_card = Cards.objects.get(card_number=to_account_number)
            
            if from_card.money >= amount:
                # Realiza la transacción
                from_card.money -= amount
                to_card.money += amount
                from_card.save()
                to_card.save()

                # Crea el registro de la transacción
                Transaction.objects.create(
                    user=request.user,
                    from_account=from_card,
                    to_account=to_card,
                    amount=amount,
                )
                print(f"Create Transaction{from_card} ---------- {to_card}")
                return redirect('transaction_success')  # Redirige a una página de éxito
            else:
                messages.error(request, 'Saldo insuficiente en la cuenta de origen.')
        
        except Cards.DoesNotExist:
            messages.error(request, 'Una de las cuentas no existe o no tiene permiso para acceder a ella.')
    
    # Renderiza el formulario si no es POST o si ocurre un error
    form = TransactionForm()
    return render(request, 'costcenter/RealizarTransferencias.html', {'form': form})

def transaction_success(request):
    return render(request, 'costcenter/transaction_success.html')

def show_cards(request):
    non_cost_center_cards = Cards.objects.filter(is_costcenter=False, user=request.user)
    context = {'cards': non_cost_center_cards}
    print(f"show cards{context}")
    return render(request, 'costcenter/show_cards.html', context)

def get_user_cards_origen(request):
    if request.user.is_authenticated:
        cards = Cards.objects.filter(user=request.user)
        input_field_id_origen = request.GET.get('input_field_id_origen')
        context = {'cards': cards, 'input_field_id_origen': input_field_id_origen}
        print(f"get user cards{context}")
        return render(request, 'costcenter/card_selection_origen.html', context)
    else:
        return JsonResponse({'error': 'No autorizado'}, status=401)

def get_user_cards_destiny(request):
    if request.user.is_authenticated:
        cards = Cards.objects.filter(user=request.user)
        input_field_id_destino = request.GET.get('input_field_id_destino')
        context = {'cards': cards, 'input_field_id_destino': input_field_id_destino}
        print(f"get user cards{context}")
        return render(request, 'costcenter/card_selection_destiny.html', context)
    else:
        return JsonResponse({'error': 'No autorizado'}, status=401)
    



def test(request):
    context= {"test":"Jeronimo"}
    return render(request,"costcenter/info_cent_cos3.html",context=context)





def InformacionPorCentroDeCostosActual(request):
    unidad = request.user.userprofile.unit
    if request.method == 'POST':
        # Redirige a la nueva vista para mostrar la información del período actual
        return redirect('totales_informacion_actual')

    # Renderiza la página de selección de centro de costos para el período actual
    return render(request, "costcenter/InformacionPorCentroDeCostosActual.html", {'unidad':unidad})



def totales_informacion_actual(request):
    # Obtener la unidad del usuario logueado
    unidad = request.user.userprofile.unit

    # Establecer la fecha de inicio del mes actual
    fecha_inicio_mes_actual = timezone.now().replace(day=1)

    tarjeta_costcenter = Cards.objects.filter(user=request.user, is_costcenter=True).first()

    # Si existe una tarjeta de tipo centro de costos, tomar su número
    costcenter_number = tarjeta_costcenter.card_number if tarjeta_costcenter else "No asignado"


    # Filtrar las tarjetas que no son de centro de costos
    tarjetas = Cards.objects.filter(user=request.user, is_costcenter=False)

    datos_tarjetas = []
    total_consumos = 0
    total_creditos = 0
    total_debitos = 0
    saldo_inicial_total = 0  # Variable para la suma de los saldos iniciales

    for tarjeta in tarjetas:
        # Obtener el saldo actual de la tarjeta
        saldo_actual = tarjeta.money

        # Calcular los consumos del mes actual
        consumos_mes_actual = Consumo.objects.filter(
            card=tarjeta,
            consumo_date__gte=fecha_inicio_mes_actual,
            consumo_date__lte=timezone.now()
        ).aggregate(total=Sum('importe'))['total'] or 0.0
        total_consumos += consumos_mes_actual

        # Calcular los créditos (transferencias hacia la tarjeta) del mes actual
        creditos_mes_actual = Transaction.objects.filter(
            to_account=tarjeta,
            movement_date__gte=fecha_inicio_mes_actual,
            movement_date__lte=timezone.now()
        ).aggregate(total=Sum('amount'))['total'] or 0.0
        total_creditos += creditos_mes_actual

        # Calcular los débitos (transferencias desde la tarjeta) del mes actual
        debitos_mes_actual = Transaction.objects.filter(
            from_account=tarjeta,
            movement_date__gte=fecha_inicio_mes_actual,
            movement_date__lte=timezone.now()
        ).aggregate(total=Sum('amount'))['total'] or 0.0
        total_debitos += debitos_mes_actual

        # Calcular el saldo anterior sumando los débitos y restando los créditos al saldo actual
        saldo_anterior = saldo_actual + debitos_mes_actual + consumos_mes_actual - creditos_mes_actual
        saldo_inicial_total += saldo_anterior  # Sumar el saldo anterior al total

        # Preparar los datos de cada tarjeta
        datos_tarjetas.append({
            'tarjeta': tarjeta.card_number,
            'denominacion': tarjeta.card_name,
            'cuenta': (tarjeta.card_number// 37200000),
            'saldo_anterior': saldo_anterior,
            'credito': creditos_mes_actual,
            'debito': debitos_mes_actual + consumos_mes_actual,
            'acredit': creditos_mes_actual,
            'otros': 0,  # Puedes agregar lógica adicional si es necesario
            'extracciones': debitos_mes_actual,
            'consumos': consumos_mes_actual,
            'saldo': saldo_actual
        })

    # Pasar los datos al contexto
    context = {
        'cc_number':costcenter_number,
        'unidad': unidad,
        'periodo': fecha_inicio_mes_actual.strftime("%m-%Y"),
        'datos_tarjetas': datos_tarjetas,
        'total_consumos': total_consumos,
        'total_creditos': total_creditos,
        'total_debitos': total_debitos,
        'saldo_inicial_total': saldo_inicial_total,  # Agregar el saldo inicial total al contexto
    }

    return render(request, 'costcenter/PeriodoActual/Totales_informacion.html', context)




def InformacionYMovimientosPorTarjetasActual(request):
    unidad = request.user.userprofile.unit
    if request.method == 'POST':
        tarjeta = request.POST.get('tarjeta')
        
        if tarjeta:
            try:
                # Verificar que la tarjeta exista y pertenezca al usuario actual
                card = Cards.objects.get(card_number=tarjeta, user=request.user, is_costcenter=False)
                today = timezone.now()
                mes = today.month
                anio = today.year

                # Redirigir a `detalle_tarjeta_mes` con los parámetros de tarjeta y mes actual
                return redirect(reverse('detalle_tarjeta_mes', kwargs={'tarjeta': card.card_number, 'periodo': f"{mes}-{anio}"}))
            
            except Cards.DoesNotExist:
                messages.error(request, "Error, tarjeta inexistente")
        else:
            messages.error(request, "Debe ingresar un número de tarjeta.")
    
    return render(request, "costcenter/InformacionYMovimientosPorTarjetasActual.html",{'unidad':unidad})








def InformacionPorCentroDeCostosAnterior(request):
    unidad = request.user.userprofile.unit
    if request.method == 'POST':
        form = MesesForm(request.POST)
        if form.is_valid():
            periodo = form.cleaned_data['periodo']
            print(f"Periodo válido: {periodo}")  # Esto debería aparecer en consola si el formulario es válido
            # Redirige con el periodo usando guion
            return redirect(reverse('totales_informacion', kwargs={'periodo': periodo}))
        else:
            print("Errores en el formulario:", form.errors)  # Muestra los errores si los hay
    else:
        form = MesesForm()
    return render(request, "costcenter/InformacionPorCentroDeCostosAnterior.html", {'form': form, 'unit': unidad})





def InformacionPorCentroDeCostosAnterior_Totales(request, periodo):
    # Parsear el periodo (mes y año)
    mes, anio = map(int, periodo.split('-'))
    # Crear fecha límite para el fin del mes usando zona horaria
    fecha_limite = timezone.make_aware(datetime(anio, mes, 1))
    fecha_fin_mes = (fecha_limite + timedelta(days=31)).replace(day=1) - timedelta(days=1)

    # Obtener todas las tarjetas del usuario logueado
    tarjetas = Cards.objects.filter(user=request.user, is_costcenter=False)
    unidad = request.user.userprofile.unit
    datos_tarjetas = []

    # Verificar si el periodo solicitado es el periodo actual
    ahora = timezone.now()
    es_periodo_actual = (anio == ahora.year and mes == ahora.month)

    for tarjeta in tarjetas:
        saldo_actual = tarjeta.money  # Asumimos inicialmente que el saldo actual es el saldo de la tarjeta

        # Si estamos en un periodo anterior al actual, ajustamos el saldo actual en función de movimientos posteriores
        if not es_periodo_actual:
            # Movimientos posteriores al mes en análisis
            creditos_posteriores = Transaction.objects.filter(
                to_account=tarjeta,
                movement_date__gt=fecha_fin_mes
            ).aggregate(total_credito=Sum('amount'))['total_credito'] or 0.0

            debitos_posteriores = Transaction.objects.filter(
                from_account=tarjeta,
                movement_date__gt=fecha_fin_mes
            ).aggregate(total_debito=Sum('amount'))['total_debito'] or 0.0

            consumos_posteriores = Consumo.objects.filter(
                card=tarjeta,
                consumo_date__gt=fecha_fin_mes
            ).aggregate(total_consumo=Sum('importe'))['total_consumo'] or 0.0

            # Ajustar saldo actual para reflejar el saldo a fin de mes en el periodo anterior
            saldo_actual = tarjeta.money - creditos_posteriores + debitos_posteriores + consumos_posteriores

        # Movimientos específicos del mes en análisis
        creditos_mes_actual = Transaction.objects.filter(
            to_account=tarjeta,
            movement_date__year=anio,
            movement_date__month=mes
        ).aggregate(total_credito=Sum('amount'))['total_credito'] or 0.0

        debitos_mes_actual = Transaction.objects.filter(
            from_account=tarjeta,
            movement_date__year=anio,
            movement_date__month=mes
        ).aggregate(total_debito=Sum('amount'))['total_debito'] or 0.0

        consumos_mes_actual = Consumo.objects.filter(
            card=tarjeta,
            consumo_date__year=anio,
            consumo_date__month=mes
        ).aggregate(total_consumo=Sum('importe'))['total_consumo'] or 0.0

        # Calcula el saldo inicial en función del saldo final ajustado y los movimientos del mes
        saldo_inicial = saldo_actual + debitos_mes_actual + consumos_mes_actual - creditos_mes_actual

        # Añadir datos al diccionario para cada tarjeta
        datos_tarjetas.append({
            'tarjeta': tarjeta.card_number,
            'denominacion': tarjeta.card_name,
            'cuenta': (tarjeta.card_number // 37200000),  # Cálculo para la cuenta asociada
            'saldo_anterior': saldo_inicial,  # Saldo inicial calculado
            'credito': creditos_mes_actual,
            'debito': debitos_mes_actual,
            'otros': 0.0,  # Lógica adicional si es necesario
            'consumos': consumos_mes_actual,
            'saldo': saldo_actual,  # Saldo final (ajustado si es periodo anterior)
        })

    # Calcular el saldo inicial total y el total de consumos del mes
    saldo_inicial_total = sum(item['saldo_anterior'] for item in datos_tarjetas)
    total_consumos = sum(item['consumos'] for item in datos_tarjetas)

    # Preparar el contexto con los datos estructurados
    context = {
        'datos_tarjetas': datos_tarjetas,
        'periodo': periodo,
        'unidad': unidad,
        'saldo_inicial': saldo_inicial_total,
        'total_consumos': total_consumos,
    }
    
    return render(request, "costcenter/PeriodoAnterior/Totales_informacion.html", context)


def detalle_tarjeta_mes_actual(request, tarjeta):
    # Obtener mes y año actuales
    mes_actual = datetime.now().month
    anio_actual = datetime.now().year

    # Obtener la tarjeta específica
    card = get_object_or_404(Cards, card_number=tarjeta, user=request.user)

    # Filtrar los consumos de la tarjeta en el mes y año actuales
    consumos = Consumo.objects.filter(
        card=card,
        consumo_date__month=mes_actual,
        consumo_date__year=anio_actual
    )

    # Calcular el total de consumos en el periodo actual
    total_consumo = consumos.aggregate(total=Sum('importe'))['total'] or 0.0

    # Calcular el saldo inicial de la tarjeta restando el total de consumos del saldo actual
    saldo_inicial = card.money - total_consumo

    # Manejo del formulario para agregar detalles
    if request.method == 'POST':
        form = ConsumoForm(request.POST)
        if form.is_valid():
            nuevo_consumo = form.save(commit=False)
            nuevo_consumo.card = card
            nuevo_consumo.consumo_date = datetime.now()  # Fecha actual para el nuevo consumo
            nuevo_consumo.save()
            return redirect('detalle_tarjeta_mes_actual', tarjeta=tarjeta)
    else:
        form = ConsumoForm()

    context = {
        'card': card,
        'saldo_inicial': saldo_inicial,
        'consumos': consumos,
        'total_consumo': total_consumo,
        'form': form,
    }
    return render(request, 'costcenter/PeriodoActual/detalle_tarjeta_mes.html', context)


def agregar_detalle(request, consumo_id):
    # Obtener el consumo específico por ID
    consumo = get_object_or_404(Consumo, consumo_id=consumo_id)
    card = consumo.card  # Obtener la tarjeta asociada al consumo

    if request.method == 'POST':
        # Actualizar los detalles del consumo desde el formulario
        consumo.tipo_de_gasto = request.POST.get('tipo_de_gasto') or None
        consumo.cantidad = request.POST.get('cantidad') or None
        consumo.un_med = request.POST.get('un_med') or None
        consumo.clase = request.POST.get('clase') or None
        consumo.monto_parcial = request.POST.get('monto_parcial') or None
        consumo.nro_factura = request.POST.get('numeroFactura') or None
        consumo.nro_movil = request.POST.get('numeroMovil') or None
        consumo.odometro = request.POST.get('odometro') or None

        # Guardar los detalles del consumo
        consumo.save()

        # Redireccionar de nuevo a la página de detalle de la tarjeta actual
        return redirect('detalle_tarjeta_mes_actual', tarjeta=consumo.card.card_number)
    print(consumo.num_ticket)
    # Pasar al contexto los datos de la tarjeta y el consumo, incluyendo los datos adicionales
    context = {
        'num_ticket': consumo.num_ticket or "",
        'consumo': consumo,
        'card_number': card.card_number,
        'card_name': card.card_name,
        'saldo_actual': card.money,
        'consumo_date': consumo.consumo_date,
        'importe': consumo.importe,
        'establecimiento': consumo.establecimiento,
        'numeroFactura': consumo.nro_factura or '',
        'numeroMovil': consumo.nro_movil or '',
        'odometro': consumo.odometro or '',
    }
    
    return render(request, 'costcenter/PeriodoActual/agregar_detalle.html', context)







def detalle_tarjeta_mes(request, tarjeta, periodo):
    # Parsear el mes y año desde el periodo
    mes, anio = map(int, periodo.split('-'))
    
    # Obtener la tarjeta específica
    card = get_object_or_404(Cards, card_number=tarjeta, user=request.user)
    
    # Filtrar los consumos de la tarjeta en el mes y año dados
    consumos = Consumo.objects.filter(
        card=card,
        consumo_date__month=mes,
        consumo_date__year=anio
    )

    # Calcular el total de consumos en el periodo
    total_consumo = consumos.aggregate(total=Sum('importe'))['total'] or 0.0

    # Calcular el saldo inicial de la tarjeta al comienzo del mes seleccionado
    # Aquí asumimos que el saldo inicial es el saldo actual de la tarjeta más el total de los consumos del mes
    saldo_inicial = card.money + total_consumo

    # Pasar los datos al template
    context = {
        'card': card,
        'saldo_inicial': saldo_inicial,
        'consumos': consumos,
        'total_consumo': total_consumo,
        'periodo': periodo,
    }
    return render(request, 'costcenter/PeriodoAnterior/detalle_tarjeta_mes.html', context)



def InformacionYMovimientosPorTarjetasAnterior(request):
    form = MesesForm(request.POST or None)

    if request.method == 'POST':
        periodo = request.POST.get('periodo')
        tarjeta = request.POST.get('tarjeta')

        if periodo and tarjeta:
            if '-' in periodo:
                try:
                    mes, anio = map(int, periodo.split('-'))
                    card = Cards.objects.get(card_number=tarjeta, user=request.user)

                    # Verificar que la tarjeta no sea un centro de costos
                    if card.is_costcenter:
                        messages.error(request, "Error, tarjeta inexistente")
                    else:
                        # Redirigir a la vista `detalle_tarjeta_mes` si todo es correcto
                        return redirect(reverse('detalle_tarjeta_mes', kwargs={'tarjeta': card.card_number, 'periodo': f"{mes}-{anio}"}))

                except Cards.DoesNotExist:
                    messages.error(request, "Error, tarjeta inexistente")

                except ValueError:
                    messages.error(request, "Formato de período incorrecto.")
            else:
                messages.error(request, "El formato de período debe ser MM-AAAA.")
        else:
            messages.error(request, "Debe ingresar un período y un número de tarjeta.")
    
    return render(request, "costcenter/InformacionYMovimientosPorTarjetasAnterior.html", {'form': form})




def RealizarDistribucion(request):
    if request.method == 'POST':
        # Obtener el centro de costos
        cost_center = Cards.objects.filter(is_costcenter=True).first()
        if not cost_center:
            messages.error(request, 'No hay centro de costos disponible.')
            return redirect('realizar_distribucion')
        
        total_amount_to_transfer = 0
        transactions = []
        
        # Recoger y procesar las solicitudes de transferencia
        for key, value in request.POST.items():
            if key.isdigit() and value:
                try:
                    card_number = key
                    amount = float(value)
                    card = Cards.objects.get(card_number=card_number, user=request.user)
                    
                    if amount > 0 and amount <= cost_center.money:
                        total_amount_to_transfer += amount
                        transactions.append((card, amount))
                    else:
                        messages.error(request, f'Fondos insuficientes para transferir {amount} a la tarjeta {card.card_number}.')
                except (ValueError, Cards.DoesNotExist):
                    continue
        
        # Verificar si el monto total a transferir es válido
        if total_amount_to_transfer <= cost_center.money:
            # Realizar las transacciones y actualizar los saldos
            for card, amount in transactions:
                Distribution.objects.create(from_account=cost_center, to_account=card, amount=amount, user=request.user)
                card.money += amount
                card.save()
                cost_center.money -= amount
            cost_center.save()
        else:
            messages.error(request, 'Fondos insuficientes en el centro de costos.')

        return redirect('realizar_distribucion')

    # Obtener tarjetas de centro de costos y tarjetas no centro de costos
    cost_center_cards = Cards.objects.filter(is_costcenter=True, user=request.user)
    non_cost_center_cards = Cards.objects.filter(is_costcenter=False, user=request.user)
    
    context = {
        'costcenter': cost_center_cards,
        'cards': non_cost_center_cards,
    }
    
    return render(request, "costcenter/RealizarDistribucion.html", context)






def DistribucionesDeFondosRealizadas(request):
    context= {"test":"Jeronimo"}
    return render(request,"costcenter/DistribucionesDeFondosRealizadas.html",context=context)



def RealizarTransferencias(request):
    cards = Cards.objects.filter(user=request.user)
    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            from_card = transaction.from_account
            to_card = transaction.to_account
            amount = transaction.amount

            if from_card.money >= amount:
                from_card.money -= amount
                to_card.money += amount
                from_card.save()
                to_card.save()
                transaction.save()
                return redirect('transaction_success')  # Redirect to a success page
            else:
                return render(request, 'costcenter/RealizarTransferencias.html', {'cards': cards, 'form': form, 'error': 'Saldo insuficiente en la cuenta de origen.'})
    else:
        form = TransactionForm()

    context = {
        'cards': cards,
        'form': form,
        'test': "Jeronimo"
    }
    return render(request, 'costcenter/RealizarTransferencias.html', context)

def get_account_balance(request, card_id):
    card = get_object_or_404(Cards, id=card_id)
    data = {
        'previous_balance': card.previous_balance,  # Assuming previous_balance is a field in Cards
        'current_balance': card.money
    }
    return JsonResponse(data)



def TransferenciasRealizadas(request):
    context= {"test":"Jeronimo"}
    return render(request,"costcenter/TransferenciasRealizadas.html",context=context)

def RealizarDevoluciones(request):
    context= {"test":"Jeronimo"}
    return render(request,"costcenter/RealizarDevoluciones.html",context=context)

def DevolucionesRealizadas(request):
    context= {"test":"Jeronimo"}
    return render(request,"costcenter/DevolucionesRealizadas.html",context=context)

def AutorizacionesPorTarjetas(request):
    context= {"test":"Jeronimo"}
    return render(request,"costcenter/AutorizacionesPorTarjetas.html",context=context)




def UltimasLiquidaciones(request):
    context= {"test":"Jeronimo"}
    return render(request,"costcenter/UltimasLiquidaciones.html",context=context)



def consulta_distribuciones(request):
    distribuciones = Distribution.objects.all()

    # Filtrar distribuciones del usuario logueado
    usuario = request.user
    distribuciones = distribuciones.filter(user=usuario)

    if request.method == 'POST':
        # Obtén los valores del formulario
        numero_distribucion = request.POST.get('id')
        cuenta_destino = request.POST.get('cuentaDestino')
        fecha_desde = request.POST.get('fechaDesde')
        fecha_hasta = request.POST.get('fechaHasta')

        # Aplica los filtros según los valores recibidos
        if numero_distribucion:
            distribuciones = distribuciones.filter(id=numero_distribucion)
        if cuenta_destino:
            distribuciones = distribuciones.filter(to_account__card_number=cuenta_destino)
        
        # Convierte las fechas al formato YYYY-MM-DD para que Django las acepte
        if fecha_desde:
            try:
                fecha_desde = datetime.strptime(fecha_desde, "%d-%m-%Y").strftime("%Y-%m-%d")
            except ValueError:
                fecha_desde = None
        if fecha_hasta:
            try:
                fecha_hasta = datetime.strptime(fecha_hasta, "%d-%m-%Y").strftime("%Y-%m-%d")
            except ValueError:
                fecha_hasta = None

        # Aplica los filtros de rango de fecha
        if fecha_desde and fecha_hasta:
            distribuciones = distribuciones.filter(distribution_date__range=[fecha_desde, fecha_hasta])
        elif fecha_desde:
            distribuciones = distribuciones.filter(distribution_date__gte=fecha_desde)
        elif fecha_hasta:
            distribuciones = distribuciones.filter(distribution_date__lte=fecha_hasta)

    # Calcula el total de montos después de aplicar los filtros
    total_monto = distribuciones.aggregate(total=Sum('amount'))['total'] or 0

    # Contexto para el template
    context = {
        'distribuciones': distribuciones,
        'total_monto': total_monto,
        'usuario_id': usuario.id,
    }
    return render(request, 'costcenter/distribuciones_realizadas_filtradas.html', context)


def detalle_distribucion(request, distribucion_id):
    distribucion = get_object_or_404(Distribution, id=distribucion_id)
    context = {'distribucion': distribucion}
    return render(request, 'costcenter/detalle_distribucion.html', context)

def detalle_distribucion(request, distribucion_id):
    distribucion = get_object_or_404(Distribution, id=distribucion_id)
    context = {
        'distribucion': distribucion,
        'user': distribucion.user,
        'user_fullname': distribucion.user.userprofile.user_fullname,
        'fecha': distribucion.distribution_date.strftime('%d/%m/%Y'),
        'hora': distribucion.distribution_date.strftime('%H:%M:%S'),
        'from_account': distribucion.from_account.card_number,
        'to_account': distribucion.to_account.card_number,
        'to_account_name':distribucion.to_account.card_name,
        'amount': distribucion.amount,
    }
    return render(request, 'costcenter/distribucion_detalle.html', context)


def consulta_transferencias(request):
    transferencias = Transaction.objects.all()

    # Filtrar transferencias del usuario logueado
    usuario = request.user
    transferencias = transferencias.filter(user=usuario)

    if request.method == 'POST':
        # Obtén los valores del formulario
        numero_transferencia = request.POST.get('id')
        cuenta_destino = request.POST.get('cuentaDestino')
        fecha_desde = request.POST.get('fechaDesde')
        fecha_hasta = request.POST.get('fechaHasta')

        # Aplica los filtros según los valores recibidos
        if numero_transferencia:
            transferencias = transferencias.filter(id=numero_transferencia)
        if cuenta_destino:
            transferencias = transferencias.filter(to_account__card_number=cuenta_destino)
        
        # Convierte las fechas al formato YYYY-MM-DD para que Django las acepte
        if fecha_desde:
            try:
                fecha_desde = datetime.strptime(fecha_desde, "%d-%m-%Y").strftime("%Y-%m-%d")
            except ValueError:
                fecha_desde = None
        if fecha_hasta:
            try:
                fecha_hasta = datetime.strptime(fecha_hasta, "%d-%m-%Y").strftime("%Y-%m-%d")
            except ValueError:
                fecha_hasta = None

        # Aplica los filtros de rango de fecha
        if fecha_desde and fecha_hasta:
            transferencias = transferencias.filter(movement_date__range=[fecha_desde, fecha_hasta])
        elif fecha_desde:
            transferencias = transferencias.filter(movement_date__gte=fecha_desde)
        elif fecha_hasta:
            transferencias = transferencias.filter(movement_date__lte=fecha_hasta)

    # Calcula el total de montos después de aplicar los filtros
    total_monto = transferencias.aggregate(total=Sum('amount'))['total'] or 0

    # Contexto para el template
    context = {
        'transferencias': transferencias,
        'total_monto': total_monto,
        'usuario_id': usuario.id,
    }
    return render(request, 'costcenter/transferencias_realizadas_filtradas.html', context)

# Vista para mostrar el detalle de una transferencia específica
def detalle_transferencia(request, transferencia_id):
    transferencia = get_object_or_404(Transaction, id=transferencia_id)
    context = {
        'transferencia': transferencia,
        'user': transferencia.user,
        'user_fullname': transferencia.user.userprofile.user_fullname,
        'fecha': transferencia.movement_date.strftime('%d/%m/%Y'),
        'hora': transferencia.movement_date.strftime('%H:%M:%S'),
        'from_account': transferencia.from_account.card_number,
        'to_account': transferencia.to_account.card_number,
        'to_account_name': transferencia.to_account.card_name,
        'amount': transferencia.amount,
        'numero_transferencia': transferencia.id,  # Asegúrate de incluirlo aquí
    }
    return render(request, 'costcenter/transferencia_detalle.html', context)

# Vista para mostrar todas las transferencias realizadas
def TransferenciasDeFondosRealizadas(request):
    context = {"test": "Jeronimo"}
    return render(request, "costcenter/TransferenciasRealizadas.html", context)








@login_required
def generar_pdf_rendicion_cc(request):
    form = MesesForm(request.POST or None)
    unit = request.user.userprofile.unit
    if request.method == "POST" and form.is_valid():
        periodo = form.cleaned_data['periodo']

        try:
            mes, anio = map(int, periodo.split('-'))
        except ValueError:
            return render(request, 'costcenter/RendicionesPorCentroDeCostosPDF.html', {
                'form': form,
                'error': "Formato de período inválido."
            })

        # Obtener unidad del usuario logueado
        user_profile = UserProfile.objects.filter(user=request.user).first()
        if not user_profile:
            return render(request, 'costcenter/RendicionesPorCentroDeCostosPDF.html', {
                'form': form,
                'error': "El usuario no tiene una unidad asociada."
            })
        cuenta = user_profile.unit  # Nombre de la unidad

        # Calculamos el rango de fechas para el mes seleccionado
        inicio_mes = timezone.make_aware(datetime(anio, mes, 1))
        if mes == 12:
            fin_mes = timezone.make_aware(datetime(anio + 1, 1, 1)) - timedelta(seconds=1)
        else:
            fin_mes = timezone.make_aware(datetime(anio, mes + 1, 1)) - timedelta(seconds=1)

        # Filtrar los datos según el rango de fechas
        transacciones = Transaction.objects.filter(
            movement_date__gte=inicio_mes,
            movement_date__lte=fin_mes
        )

        distribuciones = Distribution.objects.filter(
            distribution_date__gte=inicio_mes,
            distribution_date__lte=fin_mes
        )

        acreditaciones = Acreditaciones.objects.filter(
            fecha_hora__gte=inicio_mes,
            fecha_hora__lte=fin_mes
        )

        acreditacion_origen = "46874513789485"

        # Combinar y formatear los datos en una sola lista
        movimientos = []
        total_transacciones = 0
        total_distribuciones = 0
        total_acreditaciones = 0

        # Agregar transacciones
        for transaccion in transacciones:
            movimientos.append({
                "date": transaccion.movement_date.strftime("%d/%m/%Y"),
                "time": transaccion.movement_date.strftime("%H:%M:%S"),
                "id": transaccion.id,
                "type": "TC",
                "from_account": transaccion.from_account.card_number,
                "to_account": transaccion.to_account.card_number,
                "amount": float(transaccion.amount),
            })
            total_transacciones += float(transaccion.amount)

        # Agregar distribuciones
        for distribucion in distribuciones:
            movimientos.append({
                "date": distribucion.distribution_date.strftime("%d/%m/%Y"),
                "time": distribucion.distribution_date.strftime("%H:%M:%S"),
                "id": distribucion.id,
                "type": "DC",
                "from_account": distribucion.from_account.card_number,
                "to_account": distribucion.to_account.card_number,
                "amount": float(distribucion.amount),
            })
            total_distribuciones += float(distribucion.amount)

        # Agregar acreditaciones
        for acreditacion in acreditaciones:
            movimientos.append({
                "date": acreditacion.fecha_hora.strftime("%d/%m/%Y"),
                "time": acreditacion.fecha_hora.strftime("%H:%M:%S"),
                "id": acreditacion.id,
                "type": "A",
                "from_account": acreditacion_origen,
                "to_account": acreditacion.card.card_number,
                "amount": float(acreditacion.importe),
            })
            total_acreditaciones += float(acreditacion.importe)

        # Ordenar la lista por fecha y hora
        movimientos = sorted(movimientos, key=lambda x: (x["date"], x["time"]))

        # Generar PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Rendicion_{periodo}.pdf"'

        # Crear documento PDF
        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []

        # Crear estilo personalizado para el encabezado
        header_style = ParagraphStyle(
            name="CustomHeader",
            fontName="Helvetica",  # Arial no está incluido en reportlab; usamos Helvetica como alternativa
            fontSize=11,
            alignment=TA_JUSTIFY,
            leading=14  # Espaciado entre líneas
        )

        # Encabezado personalizado
        header_text = f"VISA FLOTA. REPORTE de RENDICION POR CENTRO DE COSTOS DEL PERIODO {periodo.upper()} PARA LA CUENTA: {cuenta}"
        header = Paragraph(header_text, header_style)
        elements.append(header)
        elements.append(Spacer(1, 10))  # Espaciado reducido

        # Tabla de movimientos
        data = [["Fecha", "Hora", "Nro Transacción", "Tip Mov", "Origen", "Destino", "Importe"]]
        for movimiento in movimientos:
            data.append([
                movimiento["date"],
                movimiento["time"],
                movimiento["id"],
                movimiento["type"],
                movimiento["from_account"],
                movimiento["to_account"],
                f"${movimiento['amount']:.2f}",
            ])

        table = Table(data, colWidths=[70, 50, 90, 60, 130, 130, 70])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),  # Tamaño de letra reducido
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),  # Espaciado reducido en las filas
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))  # Espaciado entre tablas

        # Tabla de totales
        total_data = [
            ["Total Transferencias", "Total Distribuciones", "Total Acreditaciones", "Total Otros"],
            [f"${total_transacciones:.2f}", f"${total_distribuciones:.2f}", f"${total_acreditaciones:.2f}", "$0.00"]
        ]
        total_table = Table(total_data, colWidths=[120, 120, 120, 120])
        total_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(total_table)

        # Construir PDF
        doc.build(elements)
        return response

    return render(request, 'costcenter/RendicionesPorCentroDeCostosPDF.html', {'form': form, 'unit':unit})

import io
import xlsxwriter
from datetime import datetime, timedelta
from django.shortcuts import render
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from .forms import MesesForm
from costcenter.models import Transaction, Distribution, Acreditaciones, UserProfile

@login_required
def generar_xls_rendicion_cc(request):
    form = MesesForm(request.POST or None)
    unit = request.user.userprofile.unit
    if request.method == "POST" and form.is_valid():
        periodo = form.cleaned_data['periodo']

        try:
            mes, anio = map(int, periodo.split('-'))
        except ValueError:
            return render(request, 'costcenter/RendicionesPorCentroDeCostosXLSX.html', {
                'form': form,
                'error': "Formato de período inválido."
            })

        # Obtener unidad del usuario logueado
        user_profile = UserProfile.objects.filter(user=request.user).first()
        if not user_profile:
            return render(request, 'costcenter/RendicionesPorCentroDeCostosXLSX.html', {
                'form': form,
                'error': "El usuario no tiene una unidad asociada."
            })
        cuenta = user_profile.unit

        # Calcular rango de fechas
        inicio_mes = timezone.make_aware(datetime(anio, mes, 1))
        if mes == 12:
            fin_mes = timezone.make_aware(datetime(anio + 1, 1, 1)) - timedelta(seconds=1)
        else:
            fin_mes = timezone.make_aware(datetime(anio, mes + 1, 1)) - timedelta(seconds=1)

        # Filtrar datos
        transacciones = Transaction.objects.filter(
            movement_date__gte=inicio_mes,
            movement_date__lte=fin_mes
        )
        distribuciones = Distribution.objects.filter(
            distribution_date__gte=inicio_mes,
            distribution_date__lte=fin_mes
        )
        acreditaciones = Acreditaciones.objects.filter(
            fecha_hora__gte=inicio_mes,
            fecha_hora__lte=fin_mes
        )

        acreditacion_origen = "46874513789485"

        # Crear lista de movimientos
        movimientos = []
        for transaccion in transacciones:
            movimientos.append([
                transaccion.movement_date.strftime("%d/%m/%Y"),
                transaccion.movement_date.strftime("%H:%M:%S"),
                str(transaccion.id),  # Convertir a texto
                "TC",
                str(transaccion.from_account.card_number),  # Convertir a texto
                str(transaccion.to_account.card_number),  # Convertir a texto
                float(transaccion.amount),
            ])
        for distribucion in distribuciones:
            movimientos.append([
                distribucion.distribution_date.strftime("%d/%m/%Y"),
                distribucion.distribution_date.strftime("%H:%M:%S"),
                str(distribucion.id),  # Convertir a texto
                "DC",
                str(distribucion.from_account.card_number),  # Convertir a texto
                str(distribucion.to_account.card_number),  # Convertir a texto
                float(distribucion.amount),
            ])
        for acreditacion in acreditaciones:
            movimientos.append([
                acreditacion.fecha_hora.strftime("%d/%m/%Y"),
                acreditacion.fecha_hora.strftime("%H:%M:%S"),
                str(acreditacion.id),  # Convertir a texto
                "A",
                acreditacion_origen,  # Ya es texto
                str(acreditacion.card.card_number),  # Convertir a texto
                float(acreditacion.importe),
            ])

        # Ordenar movimientos por fecha y hora
        movimientos = sorted(movimientos, key=lambda x: (x[0], x[1]))

        # Crear archivo Excel
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        # Encabezado
        headers = ["Fecha", "Hora", "Nro Transacción", "Tipo Movimiento", "Origen", "Destino", "Importe"]
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header)

        # Datos
        for row_num, movimiento in enumerate(movimientos, start=1):
            for col_num, value in enumerate(movimiento):
                if col_num in [2, 4, 5]:  # Columnas de texto
                    worksheet.write_string(row_num, col_num, value)  # Forzar formato de texto
                else:
                    worksheet.write(row_num, col_num, value)

        workbook.close()

        # Respuesta
        output.seek(0)
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Rendicion_{periodo}.xlsx"'
        return response

    return render(request, 'costcenter/RendicionesPorCentroDeCostosXLSX.html', {'form': form,'unit':unit})











from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from datetime import datetime, timedelta
from django.shortcuts import render
from django.utils import timezone
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from costcenter.models import Cards, Consumo, UserProfile
from .forms import MesesForm


@login_required
def MovimientosPorTarjetasPDF(request):
    form = MesesForm(request.POST or None)
    unit = request.user.userprofile.unit

    if request.method == "POST" and form.is_valid():
        periodo = form.cleaned_data['periodo']

        try:
            mes, anio = map(int, periodo.split('-'))
        except ValueError:
            return render(request, 'costcenter/MovimientosPorTarjetasPDF.html', {
                'form': form,
                'error': "Formato de período inválido."
            })

        # Obtener unidad del usuario logueado
        user_profile = UserProfile.objects.filter(user=request.user).first()
        if not user_profile:
            return render(request, 'costcenter/MovimientosPorTarjetasPDF.html', {
                'form': form,
                'error': "El usuario no tiene una unidad asociada."
            })

        # Calcular rango de fechas
        inicio_mes = timezone.make_aware(datetime(anio, mes, 1))
        if mes == 12:
            fin_mes = timezone.make_aware(datetime(anio + 1, 1, 1)) - timedelta(seconds=1)
        else:
            fin_mes = timezone.make_aware(datetime(anio, mes + 1, 1)) - timedelta(seconds=1)

        # Obtener tarjetas del usuario (excluyendo is_costcenter=True)
        tarjetas = Cards.objects.filter(user=request.user, is_costcenter=False)

        # Traer todos los consumos asociados a las tarjetas y el período
        consumos = Consumo.objects.filter(
            card__in=tarjetas,
            consumo_date__gte=inicio_mes,
            consumo_date__lte=fin_mes
        )

        # Calcular el total de consumos
        total_consumos = consumos.aggregate(total=Sum('importe'))['total'] or 0

        # Crear documento PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="MovimientosPorTarjetas_{periodo}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Título principal
        title_style = ParagraphStyle(
            name="Title",
            fontName="Helvetica",
            fontSize=12,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        title = Paragraph(
            f"RESUMEN DE MOVIMIENTOS DE LAS TARJETAS - PERIODO {periodo.upper()}",
            title_style
        )
        elements.append(title)

        # Tabla con las tarjetas
        tarjetas_data = [["Tarjeta", "Nombre de la Tarjeta"]]
        for tarjeta in tarjetas:
            tarjetas_data.append([tarjeta.card_number, tarjeta.card_name])

        tarjetas_table = Table(tarjetas_data, colWidths=[200, 200])  # Mitad izquierda de la página
        tarjetas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(tarjetas_table)
        elements.append(Spacer(1, 20))

        # Tabla de total de consumos
        total_data = [["Total Consumos"], [f"${total_consumos:,.2f}"]]
        total_table = Table(total_data, colWidths=[400])  # Tamaño de tabla centrado
        total_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(total_table)
        elements.append(PageBreak())  # Salto de página para los movimientos por tarjeta

        # Generar movimientos por tarjeta
        for tarjeta in tarjetas:
            # Agregar un subtítulo para cada tarjeta
            subtitle = Paragraph(
                f"Movimientos de la Tarjeta: {tarjeta.card_number} ({tarjeta.card_name})",
                styles['Heading2']
            )
            elements.append(subtitle)
            elements.append(Spacer(1, 10))

            # Filtrar consumos de la tarjeta
            consumos_tarjeta = consumos.filter(card=tarjeta)

            # Tabla de consumos
            consumos_data = [["Fecha", "Autorización", "Establecimiento", "Rubro", "Importe", "Tip Mov"]]
            for consumo in consumos_tarjeta:
                consumos_data.append([
                    consumo.consumo_date.strftime("%d/%m/%Y"),  # Solo fecha
                    consumo.consumo_id,  # ID del consumo como autorización
                    consumo.establecimiento,
                    consumo.rubro,
                    f"${consumo.importe:.2f}",
                    "CONSUMO"  # Tip Mov fijo
                ])

            consumos_table = Table(consumos_data, colWidths=[60, 80, 150, 100, 80, 60])
            consumos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
            ]))
            elements.append(consumos_table)

            # Tabla de total consumo por tarjeta
            total_tarjeta = consumos_tarjeta.aggregate(total=Sum('importe'))['total'] or 0
            total_data = [["Total Consumo"], [f"${total_tarjeta:,.2f}"]]
            total_table = Table(total_data, colWidths=[400])  # Centrado debajo de los consumos
            total_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
            ]))
            elements.append(total_table)
            elements.append(PageBreak())  # Salto de página para la siguiente tarjeta

        # Construir PDF
        doc.build(elements)
        return response

    return render(request, 'costcenter/MovimientosPorTarjetasPDF.html', {'form': form,'unit':unit})



import xlsxwriter
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from costcenter.models import Cards, Consumo, UserProfile
from .forms import MesesForm


@login_required
def MovimientosPorTarjetasXLSX(request):
    form = MesesForm(request.POST or None)
    unit = request.user.userprofile.unit
    if request.method == "POST" and form.is_valid():
        periodo = form.cleaned_data['periodo']

        try:
            mes, anio = map(int, periodo.split('-'))
        except ValueError:
            return render(request, 'costcenter/MovimientosPorTarjetasXLSX.html', {
                'form': form,
                'error': "Formato de período inválido."
            })

        # Obtener unidad del usuario logueado
        user_profile = UserProfile.objects.filter(user=request.user).first()
        if not user_profile:
            return render(request, 'costcenter/MovimientosPorTarjetasXLSX.html', {
                'form': form,
                'error': "El usuario no tiene una unidad asociada."
            })

        unidad = user_profile.unit  # Nombre de la unidad

        # Calcular rango de fechas
        inicio_mes = timezone.make_aware(datetime(anio, mes, 1))
        if mes == 12:
            fin_mes = timezone.make_aware(datetime(anio + 1, 1, 1)) - timedelta(seconds=1)
        else:
            fin_mes = timezone.make_aware(datetime(anio, mes + 1, 1)) - timedelta(seconds=1)

        # Obtener tarjetas del usuario (excluyendo is_costcenter=True)
        tarjetas = Cards.objects.filter(user=request.user, is_costcenter=False)

        # Traer todos los consumos asociados a las tarjetas y el período
        consumos = Consumo.objects.filter(
            card__in=tarjetas,
            consumo_date__gte=inicio_mes,
            consumo_date__lte=fin_mes
        ).order_by('card__card_number', 'consumo_date')

        # Crear archivo Excel en memoria
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="MovimientosPorTarjetas_{periodo}.xlsx"'

        workbook = xlsxwriter.Workbook(response, {'in_memory': True})
        worksheet = workbook.add_worksheet('Movimientos')

        # Formatos para el Excel
        bold = workbook.add_format({'bold': True})
        money_format = workbook.add_format({'num_format': '$#,##0.00'})
        date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
        text_format = workbook.add_format({'num_format': '@'})  # Formato de texto

        # Escribir encabezados
        headers = [
            "Detalle CC", "Cuenta", "Nro de Tarjeta", "Denominación",
            "Fecha", "Autorización", "Establecimiento", "Rubro",
            "Importe", "Tipo de Movimiento"
        ]
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, bold)

        # Agregar datos
        row = 1
        for consumo in consumos:
            worksheet.write(row, 0, unidad)  # Detalle CC
            worksheet.write(row, 1, str(consumo.card.card_number // 37200000))  # Cuenta como string
            worksheet.write(row, 2, str(consumo.card.card_number))  # Nro de Tarjeta como string
            worksheet.write(row, 3, consumo.card.card_name)  # Denominación

            # Eliminar la zona horaria de la fecha
            fecha_sin_tz = consumo.consumo_date.replace(tzinfo=None)
            worksheet.write_datetime(row, 4, fecha_sin_tz, date_format)  # Fecha

            worksheet.write(row, 5, str(consumo.consumo_id))  # Autorización como string
            worksheet.write(row, 6, consumo.establecimiento)  # Establecimiento
            worksheet.write(row, 7, consumo.rubro)  # Rubro
            worksheet.write_number(row, 8, consumo.importe, money_format)  # Importe
            worksheet.write(row, 9, "CONSUMO EN PESOS")  # Tipo de Movimiento
            row += 1

        workbook.close()
        return response

    return render(request, 'costcenter/MovimientosPorTarjetasXLSX.html', {'form': form,'unit':unit})




from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from datetime import datetime, timedelta
from django.shortcuts import render
from django.utils import timezone
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from costcenter.models import Cards, Transaction, Distribution, UserProfile
from .forms import MesesForm


@login_required
def RendicionPorCuentaPDF(request):
    form = MesesForm(request.POST or None)
    unit = request.user.userprofile.unit
    if request.method == "POST" and form.is_valid():
        periodo = form.cleaned_data['periodo']

        try:
            mes, anio = map(int, periodo.split('-'))
        except ValueError:
            return render(request, 'costcenter/RendicionPorCuentaPDF.html', {
                'form': form,
                'error': "Formato de período inválido."
            })

        # Obtener unidad del usuario logueado
        user_profile = UserProfile.objects.filter(user=request.user).first()
        if not user_profile:
            return render(request, 'costcenter/RendicionPorCuentaPDF.html', {
                'form': form,
                'error': "El usuario no tiene una unidad asociada."
            })

        # Calcular rango de fechas
        inicio_mes = timezone.make_aware(datetime(anio, mes, 1))
        if mes == 12:
            fin_mes = timezone.make_aware(datetime(anio + 1, 1, 1)) - timedelta(seconds=1)
        else:
            fin_mes = timezone.make_aware(datetime(anio, mes + 1, 1)) - timedelta(seconds=1)

        # Obtener tarjetas asociadas al usuario (excluyendo is_costcenter=True)
        tarjetas = Cards.objects.filter(user=request.user, is_costcenter=False)

        # Obtener distribuciones y transferencias asociadas
        distribuciones = Distribution.objects.filter(
            distribution_date__gte=inicio_mes,
            distribution_date__lte=fin_mes
        ).filter(
            from_account__in=tarjetas
        ) | Distribution.objects.filter(
            distribution_date__gte=inicio_mes,
            distribution_date__lte=fin_mes
        ).filter(
            to_account__in=tarjetas
        )

        transferencias = Transaction.objects.filter(
            movement_date__gte=inicio_mes,
            movement_date__lte=fin_mes
        ).filter(
            from_account__in=tarjetas
        ) | Transaction.objects.filter(
            movement_date__gte=inicio_mes,
            movement_date__lte=fin_mes
        ).filter(
            to_account__in=tarjetas
        )

        # Crear documento PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="RendicionPorCuenta_{periodo}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Encabezado principal
        title_style = ParagraphStyle(
            name="Title",
            fontName="Helvetica",
            fontSize=14,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        title = Paragraph(
            f"Rendición por Cuenta - Período {periodo.upper()}",
            title_style
        )
        elements.append(title)

        # Tabla de tarjetas
        tarjeta_data = [["Número de Tarjeta", "Nombre"]]
        for tarjeta in tarjetas:
            tarjeta_data.append([tarjeta.card_number, tarjeta.card_name])

        tarjeta_table = Table(tarjeta_data, colWidths=[200, 300])
        tarjeta_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(tarjeta_table)
        elements.append(Spacer(1, 20))

        # Movimientos por tarjeta
        for tarjeta in tarjetas:
            elements.append(PageBreak())
            subtitle = Paragraph(
                f"Movimientos de la Tarjeta: {tarjeta.card_number} ({tarjeta.card_name})",
                styles['Heading2']
            )
            elements.append(subtitle)
            elements.append(Spacer(1, 10))

            movimientos_data = [["Fecha", "Hora", "Nro Transac", "Tipo Mov", "Origen", "Destino", "Importe"]]

            # Totales para esta tarjeta
            total_transf_credito = 0
            total_dist_credito = 0
            total_dist_debito = 0

            # Agregar distribuciones
            distribuciones_tarjeta = distribuciones.filter(from_account=tarjeta) | distribuciones.filter(
                to_account=tarjeta
            )
            for distribucion in distribuciones_tarjeta:
                tipo_mov = "Distribución Débito" if distribucion.from_account == tarjeta else "Distribución Crédito"
                importe = distribucion.amount
                if tipo_mov == "Distribución Débito":
                    total_dist_debito += importe
                else:
                    total_dist_credito += importe
                movimientos_data.append([
                    distribucion.distribution_date.strftime("%d/%m/%Y"),
                    distribucion.distribution_date.strftime("%H:%M:%S"),
                    distribucion.id,
                    tipo_mov,
                    distribucion.from_account.card_number,
                    distribucion.to_account.card_number,
                    f"${importe:.2f}"
                ])

            # Agregar transferencias
            transferencias_tarjeta = transferencias.filter(from_account=tarjeta) | transferencias.filter(
                to_account=tarjeta
            )
            for transferencia in transferencias_tarjeta:
                tipo_mov = "Transferencia Débito" if transferencia.from_account == tarjeta else "Transferencia Crédito"
                importe = transferencia.amount
                if tipo_mov == "Transferencia Crédito":
                    total_transf_credito += importe
                movimientos_data.append([
                    transferencia.movement_date.strftime("%d/%m/%Y"),
                    transferencia.movement_date.strftime("%H:%M:%S"),
                    transferencia.id,
                    tipo_mov,
                    transferencia.from_account.card_number,
                    transferencia.to_account.card_number,
                    f"${importe:.2f}"
                ])

            movimientos_table = Table(
                movimientos_data,
                colWidths=[75, 45, 95, 100, 95, 95, 65]  # Ajuste del ancho de columnas
            )
            movimientos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
            ]))
            elements.append(movimientos_table)

            # Tabla de totales en formato horizontal
            total_data_horizontal = [
                ["Total Transf Cred", "Total Dist Cred", "Total Dist Deb", "Total Otros"],
                [f"${total_transf_credito:.2f}", f"${total_dist_credito:.2f}", f"${total_dist_debito:.2f}", "$0.00"]
            ]
            total_table_horizontal = Table(total_data_horizontal, colWidths=[100, 100, 100, 100])
            total_table_horizontal.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
            ]))
            elements.append(Spacer(1, 20))
            elements.append(total_table_horizontal)

        # Construir PDF
        doc.build(elements)
        return response

    return render(request, 'costcenter/RendicionPorCuentaPDF.html', {'form': form,'unit':unit})




from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
from openpyxl import Workbook
from costcenter.models import Cards, Transaction, Distribution, UserProfile
from .forms import MesesForm
from django.contrib.auth.decorators import login_required

@login_required
def RendicionPorCuentaXLSX(request):
    form = MesesForm(request.POST or None)
    unit = request.user.userprofile.unit
    if request.method == "POST" and form.is_valid():
        periodo = form.cleaned_data['periodo']

        try:
            mes, anio = map(int, periodo.split('-'))
        except ValueError:
            return render(request, 'costcenter/RendicionPorCuentaXLSX.html', {
                'form': form,
                'error': "Formato de período inválido."
            })

        # Obtener unidad del usuario logueado
        user_profile = UserProfile.objects.filter(user=request.user).first()
        if not user_profile:
            return render(request, 'costcenter/RendicionPorCuentaXLSX.html', {
                'form': form,
                'error': "El usuario no tiene una unidad asociada."
            })

        # Calcular rango de fechas
        inicio_mes = timezone.make_aware(datetime(anio, mes, 1))
        if mes == 12:
            fin_mes = timezone.make_aware(datetime(anio + 1, 1, 1)) - timedelta(seconds=1)
        else:
            fin_mes = timezone.make_aware(datetime(anio, mes + 1, 1)) - timedelta(seconds=1)

        # Obtener tarjetas asociadas al usuario
        tarjetas = Cards.objects.filter(user=request.user, is_costcenter=False)

        # Obtener transferencias y distribuciones asociadas
        transferencias = Transaction.objects.filter(
            movement_date__gte=inicio_mes,
            movement_date__lte=fin_mes,
            from_account__in=tarjetas
        ) | Transaction.objects.filter(
            movement_date__gte=inicio_mes,
            movement_date__lte=fin_mes,
            to_account__in=tarjetas
        )

        distribuciones = Distribution.objects.filter(
            distribution_date__gte=inicio_mes,
            distribution_date__lte=fin_mes,
            from_account__in=tarjetas
        ) | Distribution.objects.filter(
            distribution_date__gte=inicio_mes,
            distribution_date__lte=fin_mes,
            to_account__in=tarjetas
        )

        # Crear archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="RendicionPorCuenta_{periodo}.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Rendición por Cuenta"

        # Agregar encabezados
        headers = [
            "Fecha", "Hora", "Nro Transac", "Tipo Mov", "Origen", "Destino", "Importe"
        ]
        ws.append(headers)

        # Agregar datos
        for tarjeta in tarjetas:
            # Transferencias
            for transferencia in transferencias.filter(from_account=tarjeta) | transferencias.filter(to_account=tarjeta):
                tipo_mov = "Transferencia Crédito" if transferencia.to_account == tarjeta else "Transferencia Débito"
            ws.append([
                transferencia.movement_date.strftime("%d/%m/%Y"),  # Fecha
                transferencia.movement_date.strftime("%H:%M:%S"),  # Hora
                str(transferencia.id),  # Nro Transac como string
                tipo_mov,  # Tipo Mov
                str(transferencia.from_account.card_number),  # Origen como string
                str(transferencia.to_account.card_number),  # Destino como string
                f"${transferencia.amount:.2f}",  # Importe
            ])

            # Distribuciones
            for distribucion in distribuciones.filter(from_account=tarjeta) | distribuciones.filter(to_account=tarjeta):
                tipo_mov = "Distribución Crédito" if distribucion.to_account == tarjeta else "Distribución Débito"
            ws.append([
                distribucion.distribution_date.strftime("%d/%m/%Y"),  # Fecha
                distribucion.distribution_date.strftime("%H:%M:%S"),  # Hora
                str(distribucion.id),  # Nro Transac como string
                tipo_mov,  # Tipo Mov
                str(distribucion.from_account.card_number),  # Origen como string
                str(distribucion.to_account.card_number),  # Destino como string
                f"${distribucion.amount:.2f}",  # Importe
            ])


        # Guardar archivo
        wb.save(response)
        return response

    return render(request, 'costcenter/RendicionPorCuentaXLSX.html', {'form': form, 'unit':unit})

